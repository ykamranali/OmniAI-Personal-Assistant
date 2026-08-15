import os
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl import Workbook, load_workbook

class ExcelService:
    """
    Service for CRUD operations on Microsoft Excel spreadsheets (.xlsx).
    """

    @staticmethod
    def create_spreadsheet(
        filepath: str,
        headers: List[str],
        data: List[List[Any]],
        sheet_name: str = "Sheet1"
    ) -> str:
        """
        Creates a new Excel spreadsheet with specified headers and data rows.
        """
        os.makedirs(os.path.dirname(os.path.abspath(filepath)), exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        if headers:
            ws.append(headers)

        for row in data:
            ws.append(row)

        wb.save(filepath)
        return os.path.abspath(filepath)

    @staticmethod
    def read_spreadsheet(filepath: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Reads an Excel spreadsheet and returns headers and rows as structured dictionaries.
        """
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Excel file not found at: {filepath}")

        wb = load_workbook(filepath, data_only=True)
        target_sheet = sheet_name if sheet_name and sheet_name in wb.sheetnames else wb.sheetnames[0]
        ws = wb[target_sheet]

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {"sheet_name": target_sheet, "headers": [], "data": []}

        headers = [str(cell) if cell is not None else "" for cell in rows[0]]
        data_rows = []

        for row in rows[1:]:
            row_dict = {}
            for idx, cell_value in enumerate(row):
                header_name = headers[idx] if idx < len(headers) and headers[idx] else f"Column_{idx+1}"
                row_dict[header_name] = cell_value
            data_rows.append(row_dict)

        return {
            "sheet_name": target_sheet,
            "sheets_available": wb.sheetnames,
            "headers": headers,
            "data": data_rows,
            "total_rows": len(data_rows)
        }

    @staticmethod
    def append_rows(filepath: str, rows: List[List[Any]], sheet_name: Optional[str] = None) -> str:
        """
        Appends data rows to an existing Excel spreadsheet.
        """
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Excel file not found at: {filepath}")

        wb = load_workbook(filepath)
        target_sheet = sheet_name if sheet_name and sheet_name in wb.sheetnames else wb.sheetnames[0]
        ws = wb[target_sheet]

        for row in rows:
            ws.append(row)

        wb.save(filepath)
        return os.path.abspath(filepath)

    @staticmethod
    def update_cell(filepath: str, cell_address: str, value: Any, sheet_name: Optional[str] = None) -> str:
        """
        Updates a specific cell (e.g. 'A1', 'B3') in an Excel spreadsheet.
        """
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Excel file not found at: {filepath}")

        wb = load_workbook(filepath)
        target_sheet = sheet_name if sheet_name and sheet_name in wb.sheetnames else wb.sheetnames[0]
        ws = wb[target_sheet]

        ws[cell_address] = value
        wb.save(filepath)
        return os.path.abspath(filepath)

excel_service = ExcelService()
